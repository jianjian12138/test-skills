#!/usr/bin/env python3
"""qa-report: 读取执行结果，生成汇总测试报告（Markdown）。

输入（可同时给）：
  --excel  <cases.xlsx>   功能用例执行表（含「状态」列：通过/失败/阻塞/未执行）
  --json   <results.json> 接口自动化结果（qa-api-runner 产出）
  --title  <报告标题>

产出：
  <outdir>/test-report.md  含通过率、模块分布、缺陷分布、质量评估与建议

设计：让报告「读表」，避免手工统计；用例与执行结果同源（参考 OpenTest 思路）。
"""
import argparse
import json
import os
from collections import defaultdict


PASS_WORDS = {"通过", "pass", "passed", "成功"}
FAIL_WORDS = {"失败", "不通过", "fail", "failed", "未通过"}
BLOCK_WORDS = {"阻塞", "block", "blocked", "挂起"}


def norm_status(v):
    if v is None:
        return "未执行"
    s = str(v).strip()
    low = s.lower()
    if low in PASS_WORDS:
        return "通过"
    if low in FAIL_WORDS:
        return "失败"
    if low in BLOCK_WORDS:
        return "阻塞"
    if s == "" or low in {"未执行", "na", "n/a", "todo", "skip", "skipped"}:
        return "未执行"
    return s


def parse_excel(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("缺少 openpyxl：pip install openpyxl")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    module_i = idx.get("模块")
    status_i = idx.get("状态") or idx.get("实际结果")
    if status_i is None:
        raise ValueError("Excel 未找到「状态」或「实际结果」列")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        module = r[module_i] if module_i is not None else "未分类"
        status = norm_status(r[status_i])
        rows.append((module, status))
    return rows


def parse_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    rows = []
    for r in data.get("results", []):
        status = "通过" if r.get("passed") else "失败"
        rows.append(("接口自动化", status))
    return rows, data.get("summary", {})


def build_report(title, excel_rows, json_rows, json_summary):
    total_map = defaultdict(lambda: defaultdict(int))
    all_rows = excel_rows + json_rows
    for module, status in all_rows:
        total_map[module][status] += 1

    # 总计
    counts = defaultdict(int)
    for module, sm in total_map.items():
        for st, n in sm.items():
            counts[st] += n
    executed = counts.get("通过", 0) + counts.get("失败", 0) + counts.get("阻塞", 0)
    passed = counts.get("通过", 0)
    rate = (passed / executed * 100) if executed else 0

    L = [f"# 测试报告：{title}", ""]
    L.append(f"- 用例总数：{sum(counts.values())}")
    L.append(f"- 已执行：{executed}　通过：{passed}　失败：{counts.get('失败',0)}　阻塞：{counts.get('阻塞',0)}　未执行：{counts.get('未执行',0)}")
    L.append(f"- **通过率（已执行）：{rate:.1f}%**")
    if json_summary:
        L.append(f"- 接口自动化：{json_summary.get('passed',0)}/{json_summary.get('total',0)} 通过")
    L.append("")

    # 模块分布
    L.append("## 各模块通过情况")
    L.append("")
    L.append("| 模块 | 通过 | 失败 | 阻塞 | 未执行 | 通过率 |")
    L.append("|---|---|---|---|---|---|")
    for module, sm in total_map.items():
        ex = sm.get("通过",0)+sm.get("失败",0)+sm.get("阻塞",0)
        rt = (sm.get("通过",0)/ex*100) if ex else 0
        L.append(f"| {module} | {sm.get('通过',0)} | {sm.get('失败',0)} | {sm.get('阻塞',0)} | {sm.get('未执行',0)} | {rt:.0f}% |")
    L.append("")

    # 缺陷分布
    fails = [(m, sm.get("失败",0)) for m, sm in total_map.items() if sm.get("失败",0)]
    blocks = [(m, sm.get("阻塞",0)) for m, sm in total_map.items() if sm.get("阻塞",0)]
    L.append("## 缺陷 / 阻塞分布")
    L.append("")
    if fails:
        L.append("- 失败用例最多的模块（优先排查）：" + "、".join(f"{m}({n})" for m, n in sorted(fails, key=lambda x:-x[1])[:5]))
    else:
        L.append("- 无失败用例。")
    if blocks:
        L.append("- 阻塞模块：" + "、".join(f"{m}({n})" for m, n in blocks))
    L.append("")

    # 质量评估与建议
    L.append("## 质量评估与建议")
    L.append("")
    if rate >= 95 and not blocks:
        L.append("- 质量良好，通过率高且无阻塞，具备上线条件（仍需 release-check 冒烟确认）。")
    elif rate >= 80:
        L.append("- 质量基本达标，但存在失败/阻塞，建议优先修复高优先级（P0/P1）缺陷后再上线。")
    else:
        L.append("- 质量风险较高，通过率偏低，不建议上线；需集中回归并修复失败用例。")
    L.append("- 失败用例应在 Phase 4 经 `qa-bug-report` 提交缺陷并跟踪至关闭。")
    L.append("- 阻塞用例需先解决环境/数据/依赖问题，再补测。")
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser(description="生成汇总测试报告")
    ap.add_argument("--excel", help="功能用例执行表 xlsx")
    ap.add_argument("--json", help="接口自动化 results.json")
    ap.add_argument("--title", default="测试轮次")
    ap.add_argument("--outdir", default=".")
    args = ap.parse_args()

    excel_rows, json_rows, json_summary = [], [], {}
    if args.excel:
        excel_rows = parse_excel(args.excel)
    if args.json:
        json_rows, json_summary = parse_json(args.json)
    if not excel_rows and not json_rows:
        print("[error] 至少提供一个 --excel 或 --json", file=__import__("sys").stderr)
        raise SystemExit(1)

    report = build_report(args.title, excel_rows, json_rows, json_summary)
    os.makedirs(args.outdir, exist_ok=True)
    out = os.path.join(args.outdir, "test-report.md")
    with open(out, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"[ok] 测试报告已生成: {out}")


if __name__ == "__main__":
    main()
