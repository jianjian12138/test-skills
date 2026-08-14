#!/usr/bin/env python3
"""qa-test-case-gen: 从结构化用例 JSON 直接生成 Excel(.xlsx) 与 Xmind(.xmind)。

为什么用 JSON 作中间格式：测试点是半结构化文本，先由 AI 从 testpoints.md 整理成
确定性 JSON，再由本脚本稳定产出文件，避免每次重新拼装格式、保证列名/层级一致。

输入 JSON 结构（见 references/cases-schema.md）：
{
  "title": "登录模块测试用例",
  "cases": [
    {
      "id": "TC-LOGIN-001",
      "module": "登录",
      "name": "正确账号密码登录",
      "type": "正向",
      "priority": "P0",
      "precondition": "已注册用户",
      "steps": ["打开登录页", "输入正确账号", "输入正确密码", "点击登录"],
      "expected": "登录成功，跳转首页"
    }
  ]
}

用法:
    python gen_case.py --input cases.json --outdir <目录> --name <文件名基线>
产出:
    <outdir>/<name>.xlsx
    <outdir>/<name>.xmind
"""
import argparse
import json
import os
import zipfile
import uuid
import datetime


EXCEL_HEADERS = [
    "用例编号", "模块", "用例名称", "类型", "优先级",
    "前置条件", "测试步骤", "预期结果", "实际结果", "状态",
]


def load_cases(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict) or "cases" not in data:
        raise ValueError("JSON 必须包含顶层 'cases' 列表与可选 'title'")
    return data


def write_excel(data, out_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise RuntimeError("缺少依赖 openpyxl，请先安装：pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(EXCEL_HEADERS)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for c, _ in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, case in enumerate(data["cases"], 1):
        steps = case.get("steps", [])
        step_text = "\n".join(f"{n}. {s}" for n, s in enumerate(steps, 1))
        row = [
            case.get("id") or f"TC-{i:03d}",
            case.get("module", ""),
            case.get("name", ""),
            case.get("type", ""),
            case.get("priority", ""),
            case.get("precondition", ""),
            step_text,
            case.get("expected", ""),
            "",  # 实际结果（执行时填写）
            "",  # 状态（执行时填写）
        ]
        ws.append(row)

    # 列宽与换行
    widths = [12, 14, 24, 8, 8, 22, 40, 40, 14, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(out_path)


def _topic(title, note=None, children=None):
    t = {"id": uuid.uuid4().hex, "title": title}
    if note:
        t["notes"] = {"plain": {"content": note}}
    if children:
        t["children"] = {"attached": children}
    return t


def write_xmind(data, out_path):
    """构建现代 .xmind（ZIP + content.json）。"""
    modules = {}
    for case in data["cases"]:
        modules.setdefault(case.get("module", "未分类"), []).append(case)

    module_topics = []
    for mod, cases in modules.items():
        case_topics = []
        for case in cases:
            steps = case.get("steps", [])
            step_text = "\n".join(f"{n}. {s}" for n, s in enumerate(steps, 1))
            note = (
                f"类型：{case.get('type','')}\n"
                f"优先级：{case.get('priority','')}\n"
                f"前置：{case.get('precondition','')}\n"
                f"步骤：\n{step_text}\n"
                f"预期：{case.get('expected','')}"
            )
            case_topics.append(_topic(f"[{case.get('priority','')}] {case.get('name','')}", note))
        module_topics.append(_topic(mod, children=case_topics))

    sheet = {
        "id": uuid.uuid4().hex,
        "title": "测试用例",
        "rootTopic": _topic(data.get("title", "测试用例总览"), children=module_topics),
    }
    content = [sheet]
    manifest = {
        "file-entries": {"content.json": {}},
        "entries": [{"path": "content.json", "type": "json"}],
    }
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("content.json", json.dumps(content, ensure_ascii=False))
        z.writestr("metadata.json", json.dumps(
            {"creator": "qa-test-case-gen", "generated": datetime.datetime.now().isoformat()},
            ensure_ascii=False))
        z.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))


def main():
    ap = argparse.ArgumentParser(description="生成 Excel / Xmind 测试用例")
    ap.add_argument("--input", required=True, help="用例 JSON 文件")
    ap.add_argument("--outdir", required=True, help="输出目录")
    ap.add_argument("--name", default="cases", help="输出文件名基线（不含扩展名）")
    ap.add_argument("--excel-only", action="store_true")
    ap.add_argument("--xmind-only", action="store_true")
    args = ap.parse_args()

    data = load_cases(args.input)
    os.makedirs(args.outdir, exist_ok=True)

    if not args.xmind_only:
        xlsx = os.path.join(args.outdir, args.name + ".xlsx")
        write_excel(data, xlsx)
        print(f"[ok] Excel 用例: {xlsx}  ({len(data['cases'])} 条)")
    if not args.excel_only:
        xmind = os.path.join(args.outdir, args.name + ".xmind")
        write_xmind(data, xmind)
        print(f"[ok] Xmind 用例: {xmind}  ({len(data['cases'])} 条)")


if __name__ == "__main__":
    main()
